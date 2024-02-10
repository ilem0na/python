import openpyxl
from openpyxl import workbook

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

#Extracting data from the spreadsheet for products per supplier
products_per_supplier = {}
total_value_per_supplier = {}
products_less_than_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    price = product_list.cell(product_row, 3).value
    
    
    # calculate the number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
    #  print(f"Adding a new supplier {supplier_name} to the list")
        products_per_supplier[supplier_name] = 1
   
#   calculation of total value of supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier[supplier_name]
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price
    
# adding products less than 10 to a dict

    if inventory < 10:
        products_less_than_10[product_num] = int(inventory)
    
# adding values for total inventory value by supplier to the spreadsheet
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_value_per_supplier)
print(products_less_than_10)


inv_file.save("inventory_with_total_value.xlsx")